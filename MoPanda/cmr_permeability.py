import numpy as np
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
import math
import pandas as pd
from openpyxl import Workbook
import os

# Load the T2 distribution dataset from the Excel spreadsheet
input_file_path = './data/cmr/Denova1_T2_DIST_3800FT.xlsx'
df = pd.read_excel(input_file_path)

# Extract the depth column and T2 distribution data
depths = df.iloc[:, 0].values
t2_data = df.iloc[:, 1:].values

ls = np.linspace(np.log10(0.3), np.log10(3000), len(t2_data[0]))  # Generate logarithmically spaced T2 values

# Create a new Excel workbook
output_file_name = os.path.basename(input_file_path)
output_file_path = './output/' + os.path.splitext(output_file_name)[0] + 'Decomposed_Gaussian_Distributions.xlsx'
workbook = Workbook()

# Create a sheet for mean, sigma, sum, BVI, FFI, and T2lm values
peak_decomp_sheet = workbook.create_sheet(title='Peak_Decomposition_Output')

# Write the column headers for mean, sigma, sum, BVI, FFI, and T2lm
peak_decomp_sheet.cell(row=1, column=1, value='Depth')
peak_decomp_sheet.cell(row=1, column=2, value='TCMR')
peak_decomp_sheet.cell(row=1, column=3, value='BVI')
peak_decomp_sheet.cell(row=1, column=4, value='FFI')
peak_decomp_sheet.cell(row=1, column=5, value='T2lm')



# Define a function to calculate T2lm
def calculate_T2lm(t2_values, weights):
    # Calculate the logarithm of relaxation times
    log_t2 = np.log(10 ** t2_values)

    # Calculate the weighted sum of logarithms
    weighted_sum = np.sum(log_t2 * weights)

    # Calculate the sum of weights
    total_weight = np.sum(weights)

    # Calculate the geometric mean transverse relaxation time (T2lm)
    T2lm = np.exp(weighted_sum / total_weight)

    return T2lm


# Iterate through all rows in t2_data
for i in range(len(t2_data)):
    distribution = t2_data[i]

    # Calculate the sum of the T2 distribution
    tcmr = np.sum(distribution)

    # Define a model function with multiple Gaussian components
    def multi_gauss(x, *params):
        result = np.zeros_like(x)
        num_components = len(params) // 3
        for i in range(num_components):
            A, mu, sigma = params[i * 3: (i + 1) * 3]
            result += A * np.exp(-(x - mu) ** 2 / (2. * sigma ** 2))
        return result

    try:
        # Perform curve fitting with the specified number of components
        num_components = 4  # Change this to the desired number of components
        p0 = [1. / num_components, (np.log10(3000) / num_components), 1.] * num_components
        bounds = ([0., np.log10(0.3), 1e-3] * num_components, [1., np.log10(3000), 1.5] * num_components)
        maxfev = 5000  # Increase the maximum number of function evaluations
        coeff, var_matrix = curve_fit(multi_gauss, ls, distribution, p0=p0, bounds=bounds, maxfev=maxfev)

        # Generate curves for each Gaussian component
        gauss_components = []
        gauss_sums = []
        for j in range(num_components):
            pg = coeff[j * 3: (j + 1) * 3]
            gauss_components.append(multi_gauss(ls, *pg))
            gauss_sums.append(np.sum(gauss_components[-1]))

        # Calculate the BVI (Bulk Volume Irreducible) by summing the Gaussian components with 10**mean < 15
        bvi = np.sum([gauss_sums[j] for j in range(num_components) if 10 ** coeff[j * 3 + 1] < 15])

        # Calculate the FFI (Free Fluid Index) by subtracting the BVI from TCMR
        ffi = tcmr - bvi

        # If FFI is negative, set it to zero
        ffi = max(ffi, 0)

        # Subtract the Gaussian components with 10**mean < 50 from the distribution
        residual_distribution = distribution - np.sum([gauss_components[j] for j in range(num_components) if 10 ** coeff[j * 3 + 1] < 50], axis=0)

        # Set any negative values in the residual distribution to zero
        residual_distribution = np.where(residual_distribution < 0, 0, residual_distribution)

        # Calculate the T2lm using the residual distribution as the weights
        t2lm = calculate_T2lm(ls, residual_distribution)

        # Write the mean, sigma, sum, BVI, FFI, and T2lm values to the mean_sigma_sum_bvi_ffi_t2lm_sheet
        row = i + 2  # Start from row 2, since row 1 is for column headers
        peak_decomp_sheet.cell(row=row, column=1, value=depths[i])
        peak_decomp_sheet.cell(row=row, column=2, value=tcmr)
        peak_decomp_sheet.cell(row=row, column=3, value=bvi)
        peak_decomp_sheet.cell(row=row, column=4, value=ffi)
        peak_decomp_sheet.cell(row=row, column=5, value=t2lm)

        for j in range(num_components):
            column = j * 3 + 6  # Adjust the column based on the Gaussian component
            peak_decomp_sheet.cell(row=1, column=column, value=f'Mean{j + 1}')
            peak_decomp_sheet.cell(row=1, column=column + 1, value=f'Sigma{j + 1}')
            peak_decomp_sheet.cell(row=1, column=column + 2, value=f'Sum{j + 1}')
            peak_decomp_sheet.cell(row=row, column=column, value=10 ** coeff[j * 3 + 1])
            peak_decomp_sheet.cell(row=row, column=column + 1, value=coeff[j * 3 + 2])
            peak_decomp_sheet.cell(row=row, column=column + 2, value=gauss_sums[j])

    except RuntimeError:
        # If the runtime error occurs, skip the row and continue to the next iteration
        print(f"Runtime error occurred at depth {depths[i]}, row number {i + 1}")
        continue

# Remove the default sheet created by Workbook and save the Excel file
workbook.remove(workbook['Sheet'])
workbook.save(output_file_path)


# # Plot the data and fitted Gaussian components
# plt.semilogx(10 ** ls, distribution, label='Data')
# for j, component in enumerate(gauss_components):
#     plt.semilogx(10 ** ls, component, label=f'Gaussian {j + 1}')
# plt.xlabel('T2 (ms)')
# plt.ylabel('Incremental Porosity (ft³/ft³)')
# plt.title(f'Depth: {depths[i]}')
# plt.legend()
# plt.show()
