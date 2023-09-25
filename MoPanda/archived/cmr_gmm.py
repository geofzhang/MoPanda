import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from sklearn.mixture import GaussianMixture

# Load the T2 distribution dataset from the Excel spreadsheet
input_file_path = '../data/cmr/Denova1_T2_DIST_3800FT.xlsx'
df = pd.read_excel(input_file_path)

# Extract the depth column and T2 distribution data
depths = df.iloc[:, 0].values
t2_data = df.iloc[:, 1:].values
t2_values = np.linspace(np.log10(0.3), np.log10(3000), len(t2_data[0]))  # Generate logarithmically spaced T2 values
# t2_values = np.logspace(np.log10(0.3), np.log10(3000), len(t2_data[0]))  # Generate logarithmically spaced T2 values

# Perform GMM decomposition on each row
n_components = 3  # Number of components/peaks
gmm_models = []
for row in t2_data:
    data = np.concatenate([t2_values.reshape((-1, 1)), row.reshape((-1, 1))], axis=-1)
    gmm = GaussianMixture(n_components=n_components)
    gmm.fit(data)
    gmm_models.append(gmm)

# Extract the parameters of each GMM model
means = np.zeros((len(gmm_models), n_components))
height = np.zeros((len(gmm_models), n_components))
covariances = np.zeros((len(gmm_models), n_components))
weights = np.zeros((len(gmm_models), n_components))
for i, gmm in enumerate(gmm_models):
    means[i, :] = gmm.means_.flatten()[0::2]
    height[i, :] = gmm.means_.flatten()[1::2]
    covariances[i, :] = gmm.covariances_.flatten()[0::4]
    print(gmm.weights_)
    weights[i, :] = gmm.weights_.flatten()

# Create a new Excel workbook and write the results to separate sheets
output_file_name = os.path.basename(input_file_path)
output_file = './output/' + os.path.splitext(output_file_name)[0] + '_cmr_gmm_output.xlsx'
workbook = Workbook()

# Write the parameters of each peak to separate sheets
for i in range(n_components):
    sheet_name = f'Peak {i + 1}'
    sheet = workbook.create_sheet(title=sheet_name)
    result_df = pd.DataFrame(
        {'Depth': depths, 'Mean': means[:, i], 'Covariance': covariances[:, i], 'Weight': weights[:, i]})
    for row in result_df.iterrows():
        sheet.append(row[1].tolist())

# Remove the default sheet created by Workbook()
workbook.remove(workbook['Sheet'])

# Save the Excel file
workbook.save(output_file)
