import numpy as np
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
import os


class GaussianDecomposition:
    def __init__(self, df, file_name=""):
        self.df = df
        self.ls = np.linspace(np.log10(0.3), np.log10(3000), len(df.columns) - 1)
        self.file_name = file_name
        self.cutoff = 50

    def multi_gauss(self, x, *params):
        result = np.zeros_like(x)
        num_components = len(params) // 3
        for j in range(num_components):
            A, mu, sigma = params[j * 3: (j + 1) * 3]
            result += A * np.exp(-(x - mu) ** 2 / (2. * sigma ** 2))
        return result

    def decomposition_batch(self, num_components=None):
        # Extract the depth column
        depths = self.df.iloc[:, 0].values

        # Create a new Excel workbook
        output_file_path = f'./output/{os.path.splitext(self.file_name)[0]}_Decomposed_Gaussian_Distributions.xlsx'
        workbook = Workbook()

        # Create a sheet for mean, sigma, sum, BVI, FFI, T2lm, and permeability values
        peak_decomp_sheet = workbook.create_sheet(title='Peak_Decomposition_Output')

        # Write the column headers
        column_headers = ['DEPT', 'TCMR', 'BVI', 'FFI', 'T2lm_BVI', 'T2lm_FFI', 'Permeability']
        column_headers += [f'Mean{j + 1}' for j in range(num_components)]
        column_headers += [f'Sigma{j + 1}' for j in range(num_components)]
        column_headers += [f'Sum{j + 1}' for j in range(num_components)]
        for col_idx, header in enumerate(column_headers, start=1):
            peak_decomp_sheet.cell(row=1, column=col_idx, value=header)

        # Iterate through all rows in t2_data
        for row_idx, row_data in enumerate(self.df.values):
            distribution = row_data[1:]  # Exclude the first column (depth)
            tcmr = np.sum(distribution)

            try:
                # Perform curve fitting
                p0 = []
                for j in range(num_components):
                    p0.extend([1. / num_components,
                               np.log10(0.3) + j * (np.log10(3000) - np.log10(0.3)) / (num_components - 1),
                               1.])
                bounds = ([0., np.log10(0.3), 1e-3] * num_components, [1., np.log10(3000), 1.5] * num_components)
                coeff, var_matrix = curve_fit(self.multi_gauss, self.ls, distribution, p0=p0, bounds=bounds)

                # Generate curves for each Gaussian component
                gauss_components = [self.multi_gauss(self.ls, *coeff[j * 3: (j + 1) * 3]) for j in
                                    range(num_components)]
                gauss_sums = [np.sum(component) for component in gauss_components]

                # Calculate the BVI (Bulk Volume Irreducible) by summing the Gaussian components with 10**mean < 15
                bvi = np.sum([gauss_sums[j] for j in range(num_components) if 10 ** coeff[j * 3+1] < self.cutoff])

                # Calculate the FFI (Free Fluid Index) by subtracting the BVI from TCMR. If FFI is negative,
                # set it to zero.
                ffi = max(tcmr - bvi, 0)

                # Calculate the BVI distribution
                bvi_distribution = np.sum(
                    [gauss_components[j] for j in range(num_components) if 10 ** coeff[j * 3 + 1] < self.cutoff], axis=0)

                # Subtract the Gaussian components with 10**mean < 50 from the distribution
                residual_distribution = distribution - bvi_distribution

                # Set any negative values in the residual distribution to zero
                residual_distribution = np.where(residual_distribution < 0, 0, residual_distribution)

                # Calculate the T2lm_1 using the bvi distribution as the weights
                t2lm_bvi = np.exp(np.sum(np.log(10 ** self.ls) * bvi_distribution) / np.sum(bvi_distribution))

                # Calculate the T2lm_2 using the residual distribution as the weights
                t2lm_ffi = np.exp(np.sum(np.log(10 ** self.ls) * residual_distribution) / np.sum(residual_distribution))
                permeability = 10 * bvi ** 3 * t2lm_bvi + 5000 * ffi ** 6 * t2lm_ffi ** 1.5

                # Write the values to the sheet
                cell_values = [depths[row_idx], tcmr, bvi, ffi, t2lm_bvi, t2lm_ffi, permeability]
                cell_values += [10 ** coeff[j * 3 + 1] for j in range(num_components)]
                cell_values += [coeff[j * 3 + 2] for j in range(num_components)]
                cell_values += gauss_sums
                for col_idx, value in enumerate(cell_values, start=1):
                    peak_decomp_sheet.cell(row=row_idx + 2, column=col_idx, value=value)

            except RuntimeError:
                # If the runtime error occurs, skip the row and continue to the next iteration
                print(f"Runtime error occurred at depth {depths[row_idx]}, row number {row_idx + 1}")
                continue

        # Remove the default sheet created by Workbook and save the Excel file
        workbook.remove(workbook['Sheet'])
        workbook.save(output_file_path)
        print(f'Peak decomposition result saved to {output_file_path}')

        # Convert the peak_decomp_sheet to a DataFrame
        data = peak_decomp_sheet.values
        columns = next(data)[0:]
        df_results = pd.DataFrame(data, columns=columns)

        return df_results

    def decomposition_single(self, index, num_components=4, auto_search=False):
        depths = self.df.iloc[:, 0].values
        t2_data = self.df.iloc[:, 1:].values

        distribution = t2_data[index]
        tcmr = np.sum(distribution)

        # Perform automatic searching or manual mode
        if auto_search:
            min_components = 1
            max_components = 7
            aic_values = []

            for n in range(min_components, max_components + 1):
                p0 = []
                for j in range(n):
                    p0.extend([1. / n, np.log10(0.3) + j * (np.log10(3000) - np.log10(0.3)) / (n - 1), 1.])
                bounds = ([0., np.log10(0.3), 1e-3] * n, [1., np.log10(3000), 1.5] * n)
                coeff, _ = curve_fit(self.multi_gauss, self.ls, distribution, p0=p0, bounds=bounds)
                residuals = distribution - self.multi_gauss(self.ls, *coeff)
                mse = np.mean(residuals ** 2)
                num_params = 3 * n
                n_samples = len(self.ls)
                aic = n_samples * np.log(mse) + 2 * num_params
                aic_values.append(aic)

            best_n_components = np.argmin(aic_values) + min_components
            num_components = best_n_components

        # Perform curve fitting
        p0 = []
        for j in range(num_components):
            p0.extend([1. / num_components,
                       np.log10(0.3) + j * (np.log10(3000) - np.log10(0.3)) / (num_components - 1), 1.])
        bounds = ([0., np.log10(0.3), 1e-3] * num_components, [1., np.log10(3000), 1.5] * num_components)
        coeff, _ = curve_fit(self.multi_gauss, self.ls, distribution, p0=p0, bounds=bounds)
        gauss_components = [self.multi_gauss(self.ls, *coeff[j * 3: (j + 1) * 3]) for j in range(num_components)]
        gauss_sums = [np.sum(component) for component in gauss_components]
        # Calculate the BVI (Bulk Volume Irreducible) by summing the Gaussian components with 10**mean < 15
        bvi = np.sum([gauss_sums[j] for j in range(num_components) if 10 ** coeff[j * 3 + 1] < self.cutoff])

        # Calculate the FFI (Free Fluid Index) by subtracting the BVI from TCMR. If FFI is negative,
        # set it to zero.
        ffi = max(tcmr - bvi, 0)
        # Calculate the BVI distribution
        bvi_distribution = np.sum(
            [gauss_components[j] for j in range(num_components) if 10 ** coeff[j * 3 + 1] < self.cutoff], axis=0)

        # Subtract the Gaussian components with 10**mean < 50 from the distribution
        residual_distribution = distribution - bvi_distribution

        # Set any negative values in the residual distribution to zero
        residual_distribution = np.where(residual_distribution < 0, 0, residual_distribution)

        # Calculate the T2lm_1 using the bvi distribution as the weights
        t2lm_bvi = np.exp(np.sum(np.log(10 ** self.ls) * bvi_distribution) / np.sum(bvi_distribution))

        # Calculate the T2lm_2 using the residual distribution as the weights
        t2lm_ffi = np.exp(np.sum(np.log(10 ** self.ls) * residual_distribution) / np.sum(residual_distribution))
        permeability = 5 * bvi ** 4 * t2lm_bvi + 7 * ffi ** 4 * t2lm_ffi ** 2

        print('bvi:', bvi)
        print('ffi:', ffi)
        print('t2lm_bvi:', t2lm_bvi)
        print('t2lm_ffi:', t2lm_ffi)
        print('permeability:', permeability)

        # Plot the data and fitted Gaussian components
        plt.semilogx(10 ** self.ls, distribution, label='Measured')
        for j, component in enumerate(gauss_components):
            plt.semilogx(10 ** self.ls, component, label=f'Peak {j + 1}')
        plt.xlabel('T2 (ms)')
        plt.ylabel('Incremental Porosity (ft³/ft³)')
        plt.title(f'Depth: {depths[index]} ft')
        plt.legend()
        plt.show()


def perform_gaussian_decomposition(log, df, file_name, num_components=4):
    # Instantiate GaussianDecomposition class
    gaussian_decomp = GaussianDecomposition(df, file_name)

    # Call decomposition_batch method
    df_results = gaussian_decomp.decomposition_batch(num_components=num_components)

    # Calculate the number of data points in the original dataset
    N = len(df_results)

    # Calculate the current interval in the original DataFrame
    current_interval = df_results['DEPT'].diff().mean()

    df_decomp_perm = df_results.loc[:, ['DEPT', 'Permeability']]
    df_decomp_perm.set_index('DEPT', inplace=True)

    # Ensure interpolated depths end with .0 and .5
    min_depth = df_decomp_perm.index.min()
    max_depth = df_decomp_perm.index.max()
    new_index = np.arange(np.ceil(min_depth), np.floor(max_depth) + 1, 0.5)
    df_new = df_decomp_perm.reindex(new_index).interpolate(method='index')

    # Merge the interpolated DataFrame with the original DataFrame
    df_merged = pd.merge(log.df(), df_new, left_on='DEPT', right_index=True, how='left')

    log.append_curve('K_NEW', df_merged['Permeability'], unit='mD',
                     descr='Permeability Calculated from Gaussian Decomposition')
    return log
